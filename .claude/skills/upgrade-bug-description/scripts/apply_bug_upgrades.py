#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""apply_bug_upgrades.py — đọc/ghi cột Bug trong Excel cho skill upgrade-bug-description.

Usage:
  python apply_bug_upgrades.py read  <file.xlsx> [--sheet <name>]
  python apply_bug_upgrades.py write <file.xlsx> <upgrades.json>

read : tìm sheet + hàng header có cột "Bug", in JSON {sheet, header_row, bug_col, rows:[{row,id,bug}]}
write: backup file gốc vào run/backup/, thêm cột mới (mặc định "Bug (chi tiết)") NGAY BÊN PHẢI
       cột Bug (đã tồn tại thì tái sử dụng — idempotent), điền text theo số dòng, wrap text.
Cột "Bug" gốc luôn được giữ nguyên.
"""
import json
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
except ImportError:
    sys.exit("Thiếu package openpyxl. Cài: pip install openpyxl --break-system-packages")

HEADER_SCAN_ROWS = 5  # tìm header trong tối đa 5 dòng đầu


def norm(v):
    return str(v).strip().lower() if v is not None else ""


def find_bug_header(ws):
    """Trả (header_row, bug_col_idx) nếu sheet có cột 'Bug', ngược lại None."""
    for r in range(1, min(HEADER_SCAN_ROWS, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            if norm(ws.cell(row=r, column=c).value) == "bug":
                return r, c
    return None


def pick_sheet(wb, sheet_name=None):
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            sys.exit(f"Không tìm thấy sheet '{sheet_name}'. Sheet hiện có: {wb.sheetnames}")
        ws = wb[sheet_name]
        found = find_bug_header(ws)
        if not found:
            sys.exit(f"Sheet '{sheet_name}' không có cột header 'Bug' trong {HEADER_SCAN_ROWS} dòng đầu.")
        return ws, found
    for name in wb.sheetnames:
        found = find_bug_header(wb[name])
        if found:
            return wb[name], found
    sys.exit(f"Không sheet nào có cột 'Bug'. Sheet hiện có: {wb.sheetnames}")


def cmd_read(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws, (hrow, bcol) = pick_sheet(wb, sheet_name)
    id_col = None  # cột id (nếu có) để tiện trace
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(row=hrow, column=c).value) in ("id", "tc id", "stt"):
            id_col = c
            break
    rows = []
    for r in range(hrow + 1, ws.max_row + 1):
        bug = ws.cell(row=r, column=bcol).value
        rid = ws.cell(row=r, column=id_col).value if id_col else None
        if bug is None and rid is None:
            continue  # dòng hoàn toàn trống
        rows.append({"row": r, "id": rid, "bug": bug})
    print(json.dumps({
        "file": str(path), "sheet": ws.title, "header_row": hrow,
        "bug_col": bcol, "rows": rows,
    }, ensure_ascii=False, indent=2))


def backup(path):
    src = Path(path)
    bdir = Path("run/backup")
    bdir.mkdir(parents=True, exist_ok=True)
    dst = bdir / f"{src.stem}_{datetime.now():%Y%m%d_%H%M%S}{src.suffix}"
    shutil.copy2(src, dst)
    return dst


def cmd_write(path, upgrades_json):
    with open(upgrades_json, encoding="utf-8") as f:
        data = json.load(f)
    col_name = data.get("col_name") or "Bug (chi tiết)"
    upgrades = data.get("upgrades") or []
    if not upgrades:
        sys.exit("upgrades.json không có phần tử nào trong 'upgrades'.")

    bak = backup(path)
    wb = openpyxl.load_workbook(path)
    ws, (hrow, bcol) = pick_sheet(wb, data.get("sheet"))

    # Tái sử dụng cột đích nếu đã tồn tại (idempotent), ngược lại chèn ngay bên phải cột Bug
    target = None
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(row=hrow, column=c).value) == norm(col_name):
            target = c
            break
    if target is None:
        target = bcol + 1
        ws.insert_cols(target)
        head_src = ws.cell(row=hrow, column=bcol)
        head = ws.cell(row=hrow, column=target, value=col_name)
        if head_src.has_style:
            head.font = copy(head_src.font)
            head.fill = copy(head_src.fill)
            head.border = copy(head_src.border)
            head.alignment = copy(head_src.alignment)
        else:
            head.font = Font(bold=True)

    letter = ws.cell(row=hrow, column=target).column_letter
    ws.column_dimensions[letter].width = max(
        ws.column_dimensions[ws.cell(row=hrow, column=bcol).column_letter].width or 60, 60)

    written = 0
    for up in upgrades:
        r, text = int(up["row"]), up.get("text")
        if not text:
            continue
        cell = ws.cell(row=r, column=target, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        written += 1

    wb.save(path)
    print(f"OK: ghi {written} ô vào cột '{col_name}' (cột {letter}, sheet '{ws.title}') trong {path}")
    print(f"Backup file gốc: {bak}")


def main(argv):
    args = [a for a in argv if not a.startswith("--")]
    opts = {a.split("=", 1)[0].lstrip("-"): (a.split("=", 1)[1] if "=" in a else True)
            for a in argv if a.startswith("--")}
    # hỗ trợ dạng "--sheet Tên" (tách khoảng trắng)
    if opts.get("sheet") is True:
        i = argv.index("--sheet")
        if i + 1 < len(argv):
            opts["sheet"] = argv[i + 1]
            args = [a for a in args if a != argv[i + 1]]
    if len(args) < 2:
        sys.exit(__doc__)
    cmd, path = args[0], args[1]
    if not Path(path).exists():
        sys.exit(f"Không tìm thấy file: {path}")
    if cmd == "read":
        cmd_read(path, opts.get("sheet"))
    elif cmd == "write":
        if len(args) < 3:
            sys.exit("write cần thêm đường dẫn upgrades.json.\n" + __doc__)
        cmd_write(path, args[2])
    else:
        sys.exit(f"Lệnh không hợp lệ: {cmd}\n" + __doc__)


if __name__ == "__main__":
    main(sys.argv[1:])
