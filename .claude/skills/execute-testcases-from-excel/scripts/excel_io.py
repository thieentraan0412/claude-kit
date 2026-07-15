#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_io.py — Helper cho skill `execute-testcases-from-excel`.

Chức năng:
  read  : Đọc test case từ Excel (.xlsx) hoặc CSV -> JSON chuẩn hóa.
          Tự nhận diện dòng header + map cột theo alias (Tiếng Việt & English).
  write : Ghi kết quả thực thi (từ results.json) ngược vào file, thêm/điền các cột
          Actual Result / Status / Executed At / Evidence / Note-Bug ID và tô màu
          theo trạng thái (chỉ với .xlsx).

Ví dụ:
  python excel_io.py read  testcases.xlsx --sheet "Sheet1" --out run/testcases.json
  python excel_io.py write testcases.xlsx --results run/results.json --out testcases_executed.xlsx
  python excel_io.py write testcases.xlsx --results run/results.json --in-place

Ghi chú: .xls (định dạng cũ) không hỗ trợ trực tiếp — hãy lưu lại thành .xlsx.
"""
import argparse
import csv
import json
import os
import sys
import unicodedata
from datetime import datetime

# ----------------------------------------------------------------------------
# Map cột nguồn (đọc test case) — key chuẩn -> danh sách alias (đã bỏ dấu)
# ----------------------------------------------------------------------------
FIELD_ALIASES = {
    "tc_id":       ["tc id", "tcid", "tc", "ma tc", "ma test case", "ma testcase",
                    "id", "test case id", "test id", "ma", "stt tc"],
    "module":      ["module", "chuc nang", "mo dun", "phan he", "nhom chuc nang"],
    "risk_level":  ["risk level", "risk", "muc rui ro", "do rui ro", "muc do rui ro"],
    "title":       ["test title", "title", "tieu de", "ten test case", "ten testcase",
                    "test case", "mo ta", "description", "noi dung", "ten"],
    "precondition":["pre condition", "precondition", "tien dieu kien",
                    "dieu kien tien quyet", "dieu kien"],
    "steps":       ["test steps", "steps", "cac buoc", "cac buoc thuc hien",
                    "buoc thuc hien", "steps to reproduce", "thao tac", "cac buoc kiem thu"],
    "expected":    ["expected result", "expected", "ket qua mong doi",
                    "ket qua ky vong", "ket qua mong muon", "mong doi", "ket qua du kien"],
    "priority":    ["priority", "do uu tien", "muc uu tien", "uu tien"],
    "test_data":   ["test data", "du lieu test", "du lieu", "data", "du lieu kiem thu"],
}

# Cột kết quả (ghi ngược) — header chuẩn -> alias để phát hiện cột đã tồn tại
RESULT_COLUMNS = {
    "Actual Result": ["actual result", "ket qua thuc te", "actual", "thuc te"],
    "Status":        ["status", "trang thai", "ket qua chay", "verdict"],
    "Executed At":   ["executed at", "thoi gian chay", "ngay chay", "execution time"],
    "Evidence":      ["evidence", "bang chung", "screenshot", "anh", "hinh anh"],
    "Note / Bug ID": ["note / bug id", "note", "bug id", "ghi chu", "bug", "note bug id"],
}
RESULT_ORDER = list(RESULT_COLUMNS.keys())

# Màu nền theo trạng thái (ARGB) cho .xlsx
STATUS_FILL = {
    "PASS":    "FFC6EFCE",
    "FAIL":    "FFFFC7CE",
    "BLOCKED": "FFD9D9D9",
    "SKIP":    "FFFFEB9C",
}
STATUS_FONT = {
    "PASS":    "FF006100",
    "FAIL":    "FF9C0006",
    "BLOCKED": "FF3F3F3F",
    "SKIP":    "FF9C6500",
}


def norm(s):
    """Chuẩn hóa chuỗi để so khớp: thường hóa, bỏ dấu tiếng Việt, gộp khoảng trắng."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = s.replace("đ", "d")  # đ -> d (ký tự này KHÔNG tách dấu dưới NFD)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    for ch in ["_", "-", ".", ":", "(", ")", "/", "\\", "*", "#"]:
        s = s.replace(ch, " ")
    return " ".join(s.split())


def match_field(header_cell):
    """Trả về key chuẩn nếu header khớp một alias, ngược lại None."""
    h = norm(header_cell)
    if not h:
        return None
    for field, aliases in FIELD_ALIASES.items():
        if h in aliases:
            return field
    # khớp mềm: header chứa alias (cho các header dài như "Các bước thực hiện (Steps)")
    for field, aliases in FIELD_ALIASES.items():
        for a in aliases:
            if a in h and len(a) >= 3:
                return field
    return None


def detect_header(rows, scan=15):
    """Tìm dòng header: dòng trong `scan` dòng đầu map được NHIỀU cột nhất (>=2).
    Trả về (header_index, {col_index: field})."""
    best_idx, best_map = None, {}
    for i, row in enumerate(rows[:scan]):
        mapping = {}
        for j, cell in enumerate(row):
            f = match_field(cell)
            if f and f not in mapping.values():
                mapping[j] = f
        if len(mapping) > len(best_map):
            best_idx, best_map = i, mapping
    if best_idx is None or len(best_map) < 2:
        raise SystemExit(
            "[excel_io] Không nhận diện được header test case (cần >=2 cột khớp).\n"
            "  Kiểm tra file có các cột như: TC ID, Test Steps, Expected Result...\n"
            "  Hoặc chỉ định đúng --sheet."
        )
    return best_idx, best_map


# ----------------------------------------------------------------------------
# Đọc file -> danh sách các dòng (list[list[str]])
# ----------------------------------------------------------------------------
def load_rows_xlsx(path, sheet):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("[excel_io] Thiếu openpyxl. Cài: pip install openpyxl --break-system-packages")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c) for c in r])
    return rows, (sheet or wb.sheetnames[0])


def load_rows_csv(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return [list(r) for r in csv.reader(f)], "csv"


def cmd_read(args):
    ext = os.path.splitext(args.input)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        rows, sheet = load_rows_xlsx(args.input, args.sheet)
    elif ext == ".csv":
        rows, sheet = load_rows_csv(args.input)
    else:
        raise SystemExit(f"[excel_io] Định dạng không hỗ trợ: {ext} (dùng .xlsx hoặc .csv; .xls hãy lưu lại thành .xlsx).")

    header_idx, colmap = detect_header(rows)
    records = []
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        rec = {k: "" for k in FIELD_ALIASES}
        for j, field in colmap.items():
            rec[field] = (row[j].strip() if j < len(row) and row[j] is not None else "")
        if not any(str(v).strip() for v in rec.values()):
            continue  # bỏ dòng trống
        rec["_row"] = i + 1          # số dòng Excel (1-based)
        rec["_sheet"] = sheet
        records.append(rec)

    out = {
        "source": os.path.abspath(args.input),
        "sheet": sheet,
        "header_row": header_idx + 1,
        "columns_detected": {v: k for k, v in colmap.items()},
        "count": len(records),
        "testcases": records,
    }
    payload = json.dumps(out, ensure_ascii=False, indent=2)
    if args.out:
        os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(payload)
        print(f"[excel_io] Đọc {len(records)} test case -> {args.out}")
        print(f"[excel_io] Cột nhận diện: {', '.join(sorted(colmap.values()))}")
    else:
        print(payload)


# ----------------------------------------------------------------------------
# Ghi kết quả ngược vào .xlsx / .csv
# ----------------------------------------------------------------------------
def load_results(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    items = data.get("results", data) if isinstance(data, dict) else data
    by_row, by_id = {}, {}
    for it in items:
        if it.get("_row") is not None:
            by_row[int(it["_row"])] = it
        if it.get("tc_id"):
            by_id[norm(it["tc_id"])] = it
    return by_row, by_id


def result_value(item, key):
    ev = item.get("evidence")
    if key == "Actual Result":
        return item.get("actual", "")
    if key == "Status":
        return str(item.get("status", "")).strip().upper()
    if key == "Executed At":
        return item.get("executed_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if key == "Evidence":
        return ", ".join(ev) if isinstance(ev, list) else (ev or "")
    if key == "Note / Bug ID":
        note = item.get("note", "") or ""
        bug = item.get("bug_id", "") or ""
        return " ".join(x for x in [bug, note] if x).strip()
    return ""


def cmd_write(args):
    ext = os.path.splitext(args.input)[1].lower()
    by_row, by_id = load_results(args.results)
    if ext in (".xlsx", ".xlsm"):
        _write_xlsx(args, by_row, by_id)
    elif ext == ".csv":
        _write_csv(args, by_row, by_id)
    else:
        raise SystemExit(f"[excel_io] Định dạng không hỗ trợ để ghi: {ext}")


def _resolve_out(args, default_suffix="_executed"):
    if args.in_place:
        return args.input
    if args.out:
        return args.out
    stem, ext = os.path.splitext(args.input)
    return f"{stem}{default_suffix}{ext if ext in ('.xlsx', '.xlsm') else '.xlsx'}"


def _write_xlsx(args, by_row, by_id):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
    except ImportError:
        raise SystemExit("[excel_io] Thiếu openpyxl. Cài: pip install openpyxl --break-system-packages")

    wb = load_workbook(args.input)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    rows = [[("" if c.value is None else str(c.value)) for c in r] for r in ws.iter_rows()]
    header_idx, colmap = detect_header(rows)
    header_row_no = header_idx + 1

    # Tìm cột tc_id để khớp dự phòng
    tcid_col = next((j for j, f in colmap.items() if f == "tc_id"), None)

    # Xác định / tạo cột kết quả
    existing = {}
    header_cells = rows[header_idx]
    for j, cell in enumerate(header_cells):
        hn = norm(cell)
        for canon, aliases in RESULT_COLUMNS.items():
            if hn == norm(canon) or hn in aliases:
                existing[canon] = j
    next_col = len(header_cells)
    result_col = {}
    for canon in RESULT_ORDER:
        if canon in existing:
            result_col[canon] = existing[canon]
        else:
            result_col[canon] = next_col
            ws.cell(row=header_row_no, column=next_col + 1, value=canon)
            next_col += 1

    written = 0
    for i in range(header_idx + 1, len(rows)):
        excel_row = i + 1
        item = by_row.get(excel_row)
        if item is None and tcid_col is not None and tcid_col < len(rows[i]):
            item = by_id.get(norm(rows[i][tcid_col]))
        if not item:
            continue
        status = str(item.get("status", "")).strip().upper()
        for canon, jcol in result_col.items():
            ws.cell(row=excel_row, column=jcol + 1, value=result_value(item, canon))
        if status in STATUS_FILL:
            fill = PatternFill("solid", fgColor=STATUS_FILL[status])
            font = Font(color=STATUS_FONT[status], bold=(status in ("FAIL", "PASS")))
            scol = result_col["Status"] + 1
            ws.cell(row=excel_row, column=scol).fill = fill
            ws.cell(row=excel_row, column=scol).font = font
        written += 1

    out = _resolve_out(args)
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    wb.save(out)
    print(f"[excel_io] Ghi {written} kết quả -> {out}")


def _write_csv(args, by_row, by_id):
    rows, _ = load_rows_csv(args.input)
    header_idx, colmap = detect_header(rows)
    tcid_col = next((j for j, f in colmap.items() if f == "tc_id"), None)
    header = rows[header_idx]
    base = len(header)
    result_col = {c: base + k for k, c in enumerate(RESULT_ORDER)}
    header = header + RESULT_ORDER

    out_rows = rows[:header_idx] + [header]
    written = 0
    for i in range(header_idx + 1, len(rows)):
        row = list(rows[i]) + [""] * len(RESULT_ORDER)
        excel_row = i + 1
        item = by_row.get(excel_row)
        if item is None and tcid_col is not None and tcid_col < len(rows[i]):
            item = by_id.get(norm(rows[i][tcid_col]))
        if item:
            for canon, jcol in result_col.items():
                row[jcol] = result_value(item, canon)
            written += 1
        out_rows.append(row)

    stem, _ = os.path.splitext(args.input)
    out = args.input if args.in_place else (args.out or f"{stem}_executed.csv")
    with open(out, "w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(out_rows)
    print(f"[excel_io] Ghi {written} kết quả -> {out}")


def main():
    p = argparse.ArgumentParser(description="Đọc/ghi test case Excel cho skill execute-testcases-from-excel")
    sub = p.add_subparsers(dest="cmd", required=True)

    pr = sub.add_parser("read", help="Đọc Excel/CSV -> JSON test case chuẩn hóa")
    pr.add_argument("input")
    pr.add_argument("--sheet", default=None)
    pr.add_argument("--out", default=None, help="Đường dẫn JSON xuất (bỏ trống = in ra stdout)")
    pr.set_defaults(func=cmd_read)

    pw = sub.add_parser("write", help="Ghi kết quả (results.json) ngược vào file")
    pw.add_argument("input")
    pw.add_argument("--results", required=True, help="File results.json")
    pw.add_argument("--sheet", default=None)
    pw.add_argument("--out", default=None, help="File xuất (mặc định <tên>_executed.xlsx)")
    pw.add_argument("--in-place", action="store_true", help="Ghi đè trực tiếp file gốc")
    pw.set_defaults(func=cmd_write)
